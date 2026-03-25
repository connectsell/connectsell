# -*- coding: utf-8 -*-
"""
엑셀 시트를 읽어서 각 행마다 이메일 발송 후,
발송일시·발송여부를 기록해 엑셀을 저장하는 스크립트.

엑셀 컬럼: NO, 셀러명, 이메일주소, 발송일시, 발송여부
- 000님 → 셀러명으로 치환하여 발송
- 발송 성공: 발송일시 기록, 발송여부 공란
- 발송 실패: 발송일시 기록, 발송여부 "오류"
"""

import os
import sys
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl이 필요합니다. 터미널에서 실행: pip install openpyxl")
    sys.exit(1)


# ---------- 설정 (환경 변수 또는 여기서 수정) ----------
# 발신 이메일 주소 (Gmail 또는 네이버 등)
EMAIL_SENDER = os.environ.get("EMAIL_SENDER", "")
# 앱 비밀번호 (Gmail: Google 계정 → 보안 → 2단계 인증 → 앱 비밀번호)
EMAIL_PASSWORD = os.environ.get("EMAIL_APP_PASSWORD", "")

# Gmail: smtp.gmail.com, 587  /  네이버: smtp.naver.com, 587
SMTP_HOST = os.environ.get("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))

# 메일 제목
EMAIL_SUBJECT = "ConnectSell 공동구매 제안서 안내"

# HTML 템플릿 파일 (스크립트와 같은 폴더의 kakkukishake-email.html)
SCRIPT_DIR = Path(__file__).resolve().parent
HTML_TEMPLATE_PATH = SCRIPT_DIR / "kakkukishake-email.html"


def load_html_template():
    """이메일 HTML 템플릿 로드 (UTF-8)."""
    with open(HTML_TEMPLATE_PATH, "r", encoding="utf-8") as f:
        return f.read()


def make_body(html_template: str, seller_name: str) -> str:
    """000님 → 셀러명 치환."""
    return html_template.replace("000님", seller_name)


def send_email(to_address: str, html_body: str) -> None:
    """HTML 본문으로 이메일 발송 (UTF-8)."""
    if not EMAIL_SENDER or not EMAIL_PASSWORD:
        raise ValueError(
            "EMAIL_SENDER, EMAIL_APP_PASSWORD(또는 EMAIL_PASSWORD)를 설정하세요. "
            "환경 변수로 넣거나 스크립트 상단에서 수정하세요."
        )

    msg = MIMEMultipart("alternative")
    msg["Subject"] = EMAIL_SUBJECT
    msg["From"] = EMAIL_SENDER
    msg["To"] = to_address

    part = MIMEText(html_body, "html", "utf-8")
    msg.attach(part)

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
        server.starttls()
        server.login(EMAIL_SENDER, EMAIL_PASSWORD)
        server.sendmail(EMAIL_SENDER, [to_address], msg.as_string())


def find_column_index(header_row):
    """첫 행에서 컬럼 인덱스 찾기 (1-based)."""
    cols = {}
    for idx, cell in enumerate(header_row, start=1):
        if cell.value is None:
            continue
        val = str(cell.value).strip()
        if val == "NO":
            cols["no"] = idx
        elif val == "셀러명":
            cols["seller"] = idx
        elif val == "이메일주소":
            cols["email"] = idx
        elif val == "발송일시":
            cols["sent_at"] = idx
        elif val == "발송여부":
            cols["status"] = idx
    return cols


def main():
    if len(sys.argv) < 2:
        print("사용법: python send_emails.py <엑셀파일경로>")
        print("예: python send_emails.py C:\\Users\\USER\\Desktop\\셀러_이메일목록.xlsx")
        sys.exit(1)

    excel_path = Path(sys.argv[1])
    if not excel_path.exists():
        print(f"파일을 찾을 수 없습니다: {excel_path}")
        sys.exit(1)

    if not HTML_TEMPLATE_PATH.exists():
        print(f"HTML 템플릿을 찾을 수 없습니다: {HTML_TEMPLATE_PATH}")
        sys.exit(1)

    print("엑셀 로드 중...")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    header = list(ws.iter_rows(min_row=1, max_row=1))[0]
    cols = find_column_index(header)
    if not all(k in cols for k in ["seller", "email", "sent_at", "status"]):
        print("필요한 컬럼이 없습니다: 셀러명, 이메일주소, 발송일시, 발송여부")
        sys.exit(1)

    html_template = load_html_template()
    now_fmt = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    sent_ok = 0
    sent_fail = 0

    for row_num in range(2, ws.max_row + 1):
        seller = ws.cell(row=row_num, column=cols["seller"]).value
        email_addr = ws.cell(row=row_num, column=cols["email"]).value
        sent_at_cell = ws.cell(row=row_num, column=cols["sent_at"])
        status_cell = ws.cell(row=row_num, column=cols["status"])

        if not email_addr or not str(email_addr).strip():
            continue
        email_addr = str(email_addr).strip()
        seller_name = (seller or "").strip() or "크리에이터"

        # 이미 발송한 행은 건너뜀 (발송일시가 있으면 스킵)
        if sent_at_cell.value is not None and str(sent_at_cell.value).strip():
            print(f"  [{row_num}] {seller_name} ({email_addr}) - 이미 발송됨, 스킵")
            continue

        html_body = make_body(html_template, seller_name)

        try:
            send_email(email_addr, html_body)
            sent_at_cell.value = now_fmt
            status_cell.value = ""  # 공란 = 정상 발송
            sent_ok += 1
            print(f"  [{row_num}] {seller_name} ({email_addr}) - 발송 완료")
        except Exception as e:
            sent_at_cell.value = now_fmt
            status_cell.value = "오류"
            sent_fail += 1
            print(f"  [{row_num}] {seller_name} ({email_addr}) - 오류: {e}")

    wb.save(excel_path)
    print(f"\n엑셀 저장 완료: {excel_path}")
    print(f"발송 성공: {sent_ok}, 발송 실패(오류): {sent_fail}")


if __name__ == "__main__":
    main()
