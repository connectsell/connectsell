# -*- coding: utf-8 -*-
"""사용법 확인용 샘플 엑셀 생성. 실제 발송 시 본인 엑셀 파일을 사용하세요."""
import openpyxl
from pathlib import Path

path = Path(__file__).resolve().parent / "셀러_이메일목록_샘플.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "이메일목록"

# 헤더 (1행)
ws["A1"] = "NO"
ws["B1"] = "셀러명"
ws["C1"] = "이메일주소"
ws["D1"] = "발송일시"
ws["E1"] = "발송여부"
# 샘플 데이터 (2~5행) - 발송일시/발송여부는 비워 둠
for i, (no, name, email) in enumerate([
    (1, "연선", "alvl0825@naver.com"),
    (2, "호호호 공주", "skfro79@naver.com"),
    (3, "뷰티마녀", "connectsell01@gmail.com"),
    (4, "헬스남", "tendown2025@gmail.com"),
], start=2):
    ws.cell(row=i, column=1, value=no)
    ws.cell(row=i, column=2, value=name)
    ws.cell(row=i, column=3, value=email)
    # D, E열은 비움

wb.save(path)
print(f"샘플 엑셀 생성됨: {path}")
